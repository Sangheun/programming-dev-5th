import openpyxl
import re
import requests
import xmltodict


from django.forms import ValidationError
from django.utils.deconstruct import deconstructible

def lnglat_validator(lnglat):
    if not re.match(r'^(\d+\.?\d*),(\d+\.?\d*)$', lnglat):
        raise ValidationError('Invalid LngLat Type')

@deconstructible
class MinLengthValidator(object):
    def __init__(self, min_length):
        self.min_length = min_length

    def __call__(self, value):
        if len(value) < self.min_length:
            raise ValidationError('{}글자 이상 입력해 주세요.'.format(self.min_length))


'''
def min_length_validator(min_length):
    def wrap(value):
        if len(value) < min_length:
            raise ValidationError('{}글자 이상 입력해 주세요.'.format(min_length))
    return wrap

def max_length_validator(max_length):
    def wrap(value):
        if len(value) > max_length:
            raise ValidationError('{} 글자 이하로 입력해 주세요.'.format(max_length))
    return wrap
'''

def phone_number_validator(value):
    if not re.match(r'^01[06789][1-9]\d{6,7}$', value):
        raise ValidationError('올바른 휴대폰 번호를 입력해 주세요.')

# def get_zip_code(filename, sheet_name):
#     zip_code_src = openpyxl.load_workbook(filename)
#     sheet = zip_code_src.get_sheet_by_name(sheet_name)
#     get_max_row = sheet.max_row
#     zip_code = []
#     for i in range(2, get_max_row+1):
#         #첫번째 줄은 index
#         zip_code.append(sheet.cell(row=i, column=1).value)
#     return zip_code


# zipcodes = get_zip_code('practice.xlsx','Sheet1')

# def zip_code_validator(zipcode):
#     test = re.compile(r'[0]\d{4}')
#     for zipcode in zipcodes:
#         if not test.match(zipcode):
#             raise ValidationError('올바른 우편번호를 입력해주세요')




# def get_zip_code(filename, sheet_name):
#     zip_code_src = openpyxl.load_workbook(filename)
#     sheet = zip_code_src.get_sheet_by_name(sheet_name)
#     get_max_row = sheet.max_row
#     zip_code = []
#     for i in range(2, get_max_row+1):
#         #첫번째 줄은 index
#         zip_code.append(sheet.cell(row=i, column=1).value)
#     return zip_code

# zipcodes = get_zip_code('practice.xlsx', 'Sheet1')



@deconstructible
class ZipCodeValidator(object):
    def __init__(self, is_check_exist=False):
        self.is_check_exist = is_check_exist

    def __call__(self, zip_code):
        if not re.match(r'^\d{5}$', zip_code):
            raise ValidationError('5자리 숫자로 입력해주세요.')

        if self.is_check_exist:
            self.check_exist_from_db(zip_code)

    def check_exist_from_db(self, zip_code):
        from blog.models import ZipCode
        if not ZipCode.objects.filter(code=zip_code).exists():
            raise ValidationError('없는 우편번호에용')


    def check_exist(self, zip_code):

        params = {
            'regky':'5315606e3fe0e7fdf1470022382025',
            'target':'postNew',
            'query':zip_code,
        }

        xml = requests.get('http://biz.epost.go.kr/KpostPortal/openapi', params=params).text
        response = xmltodict.parse(xml)
        try:
            error = response['error']
        except KeyError:
            pass
        else:
            raise ValidationError('[{error_code}]{message}'.format(**error))







# test = re.compile(r'[0]\d{4}')

# for zipcode in zipcodes:
#     if test.match(zipcode):
#         print("Yes!")
#     else:
#         print("No!")

